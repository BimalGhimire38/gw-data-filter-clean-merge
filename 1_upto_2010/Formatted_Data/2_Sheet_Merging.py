# 2081/10/24  08:03 PM PM

import openpyxl
from openpyxl import load_workbook,Workbook


def merging_sheets(wb,well_type):
    all_sheets_list = wb.sheetnames
    new_sheet_name = 'All_'+ well_type
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = new_sheet_name
    final_file_name = f"Combined_{well_type}_upto_2010.xlsx"

    sheet_index = 0
    for sheet in all_sheets_list:
        current_district_name = sheet.split('_')[0]
        if well_type in sheet:
            ws = wb[sheet]

            # For first Sheet
            if sheet_index == 0 or new_ws.max_column <=1:
                for i in range(1,ws.max_row+1):
                    
                    if i == 1:
                        header_data = [ws.cell(row=i,column=j).value for j in range(1,ws.max_column+1)]
                        header_data[0]='District'
                        header_data.insert(1,'Station_Name')
                        new_ws.append(header_data)
                        
                       
                    # Skip the month row
                    elif i == 2:
                        continue
                    
                    else:
                        station_data = [ws.cell(row=i,column=j).value for j in range(1,ws.max_column+1)]
                        station_data.insert(0,current_district_name)
                        new_ws.append(station_data)
            else:
                for i in range(3,ws.max_row+1):
                    station_data = [ws.cell(row=i,column=j).value for j in range(1,ws.max_column+1)]
                    station_data.insert(0,current_district_name)
                    new_ws.append(station_data)
        sheet_index += 1
    new_wb.save(final_file_name)

swb = load_workbook('STW.xlsx')
# new_sws = swb.create_sheet('All_STW')
merging_sheets(wb=swb,well_type='STW')

dwb = load_workbook('DTW.xlsx')
# new_dws = dwb.create_sheet('All_DTW')
merging_sheets(wb=dwb,well_type='DTW')


