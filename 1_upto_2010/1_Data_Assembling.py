# # 2081/10/24  04:36 PM

import openpyxl
from openpyxl import load_workbook,Workbook


# def merging_sheets(wb,new_ws,well_type):
#     final_file_name = f"Combined_{well_type}_upto_2010.xlsx"

#     sheet_index = 0
#     for sheet in wb.sheetnames:
#         current_district_name = sheet.split('_')[0]
#         if 'STW' in sheet:
#             ws = wb[sheet]

#             # For first Sheet
#             if sheet_index == 0:
#                 for i in range(1,ws.max_row+1):
#                     # Skip the month row
#                     if i == 2:
#                         continue
#                     if i == 1:
#                         header_data = [ws.cell(row=i,column=j).value for j in range(1,ws.max_column+1)]
#                         header_data[0]='District'
#                         header_data.insert(1,'Station_Name')
#                         new_ws.append(header_data)
                    
#                     else:
#                         station_data = [ws.cell(row=i,column=j).value for j in range(1,ws.max_column+1)]
#                         station_data.insert(0,current_district_name)
#                         new_ws.append(station_data)
#             else:
#                 for i in range(3,ws.max_row+1):
#                     station_data = [ws.cell(row=i,column=j).value for j in range(1,ws.max_column+1)]
#                     station_data.insert(0,current_district_name)
#                     new_ws.append(station_data)
#         sheet_index += 1
#     wb.save(final_file_name)





files_list = ['Jhapa.xlsx','Morang.xlsx','Sunsari.xlsx','Udaypur.xlsx']

stwb = Workbook()
dtwb = Workbook()

for file_name in files_list:
    district_name = file_name.split('.')[0]
    wb = load_workbook(file_name)
    sheets = wb.sheetnames
    for sheet in sheets:
        ws = wb[sheet]
        # print(sheet)
        if 'STW' in str(sheet):
            # print(f"{sheet} is STW")
            new_sheet = district_name + '_STW'
            st_sheet =stwb.create_sheet(new_sheet)
            # Swap Row And Columns
            for i in range(1,ws.max_row+1):
                for j in range(1,ws.max_column+1):
                    st_sheet.cell(row=j,column=i).value = ws.cell(row=i,column=j).value
            stwb.save('STW.xlsx')
        elif 'DTW' in str(sheet):
            # print(f"{sheet} is DTW")
            new_sheet = district_name + '_DTW'
            dt_sheet =dtwb.create_sheet(new_sheet)
            # Swap Row And Columns
            for i in range(1,ws.max_row+1):
                for j in range(1,ws.max_column+1):
                    dt_sheet.cell(row=j,column=i).value = ws.cell(row=i,column=j).value
            dtwb.save('DTW.xlsx')

# swb = load_workbook('STW.xlsx')
# new_sws = swb.create_sheet('All_STW')
# merging_sheets(wb=swb,new_ws=new_sws,well_type='STW')

# dwb = load_workbook('DTW.xlsx')
# new_dws = dwb.create_sheet('All_DTW')
# merging_sheets(wb=dwb,new_ws=new_dws,well_type='DTW')


