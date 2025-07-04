# 2081/11/13 09:50 PM, @ Home, Bungamati ,Lalitpur

import openpyxl
import os
from openpyxl import load_workbook,Workbook

old_stw ='Combined_STW_upto_2010.xlsx'
old_dtw ='Combined_DTW_upto_2010.xlsx' 
new_stw ='1_STW_from_2011.xlsx'
new_dtw ='1_DTW_from_2011.xlsx'

# files = {
#     'Combined_STW_upto_2010.xlsx':'STW_formatted_from_2011.xlsx',
#     'Combined_DTW_upto_2010.xlsx':'DTW_formatted_from_2011.xlsx'
#         }
def sheet_name_searcher(current_district,sheet_list):
    for a_sheet_name in sheet_list:
            if str(current_district) in str(a_sheet_name):
                 return a_sheet_name
            else:
                 pass

# def data_inserter():
#      old_ws = 
#      new_ws = 
     
 

def Searching_Station(old,new,well_type):
    data_found_stations =[]
    header_counter =0
    old_wb = load_workbook(old)
    old_ws = old_wb.active
    target_column = old_ws.max_column+1
    new_wb = load_workbook(new)
    new_wb_sheet_list = new_wb.sheetnames
    # print(new_wb_sheet_list)
    new_ws_name = None
    for i in range(2,old_ws.max_row+1):
        current_district = str(old_ws.cell(row=i,column=1).value)
        current_station =str(old_ws.cell(row=i,column=2).value)
        for a_sheet_name in new_wb_sheet_list:
            if str(current_district) in str(a_sheet_name):
                 new_ws_name = a_sheet_name
                 break
        # new_ws_name = sheet_name_searcher(current_district=current_district,sheet_list=new_wb_sheet_list)
        new_ws = new_wb[new_ws_name]
        for a in range(1,new_ws.max_row+1):
             searched_station = str(new_ws.cell(row=a,column=2).value)
             searched_district = str(new_ws.cell(row=a,column=1).value)

             if current_district.lower() == searched_district.lower() and current_station.lower() == searched_station.lower():
                  target_row = i
                  headers_to_append =[new_ws.cell(row=1,column=k).value for k in range(5,new_ws.max_column+1)]
                  data_to_append = [new_ws.cell(row=a,column=x).value for x in range(5,new_ws.max_column+1)]
                  
                  # Appending the additional Headers
                  if header_counter == 0:
                       for index,value in enumerate(headers_to_append):
                            old_ws.cell(row=1,column=target_column+index,value=value)
                       header_counter +=1
                  # Appending Current Station's Data
                  for index,value in enumerate(data_to_append):
                       old_ws.cell(row=target_row,column=target_column+index,value=value)
                  data_found_stations.append(f"{current_district}_{current_station}")
                    

    old_wb.save(f"01_Combined_{well_type}_upto_2018.xlsx")
    print(len(data_found_stations))  
    print(data_found_stations)                     
                  
Searching_Station(old_stw,new_stw,well_type='STW')
Searching_Station(old_dtw,new_dtw,well_type='DTW')