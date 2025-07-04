# 2081/11/13 10:02 Pm, @ Home, Bungmati ,Lalitpur

import openpyxl
from openpyxl import load_workbook

def coordinate_assigner(target_sheet,reference_book,well_type):
    target_sheet.insert_cols(3,2)
    target_sheet.cell(row=1,column=3).value ="X"
    target_sheet.cell(row=1,column=4).value ="Y"
    for i in range(2,target_sheet.max_row+1):
        current_district = target_sheet.cell(row=i, column=1).value
        current_station = target_sheet.cell(row=i, column=2).value

        for data_sheet in reference_book.sheetnames:
            if data_sheet.startswith(current_district):
                reference_sheet = reference_book[data_sheet]
                for k in range(2,reference_sheet.max_row+1):
                    found_district = reference_sheet.cell(row=k, column=1).value
                    found_station = reference_sheet.cell(row=k, column=2).value
                    if str(current_station).lower() == str(found_station).lower() and str(current_district).lower() == str(found_district).lower():
                        target_sheet.cell(row=i, column=3).value = reference_sheet.cell(row=k, column=3).value
                        target_sheet.cell(row=i, column=4).value = reference_sheet.cell(row=k, column=4).value

    target_book.save(f"1_{well_type}_upto_2018_with_coordinates.xlsx")
                

# for STW
reference_book = load_workbook("1_STW_from_2011.xlsx")
target_book =load_workbook("01_Combined_STW_upto_2018.xlsx")
target_sheet = target_book.active
coordinate_assigner(target_sheet,reference_book,"STW")
# for DTW
reference_book = load_workbook("1_STW_from_2011.xlsx")
target_book =load_workbook("01_Combined_DTW_upto_2018.xlsx")
target_sheet = target_book.active
coordinate_assigner(target_sheet,reference_book,"DTW")





