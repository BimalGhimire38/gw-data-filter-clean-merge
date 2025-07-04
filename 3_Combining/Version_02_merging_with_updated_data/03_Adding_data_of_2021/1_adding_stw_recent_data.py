# 2081/11/13   11:20 PM

import openpyxl
from openpyxl.reader.excel import load_workbook


reference_book = load_workbook('1_STW_upto_2018_with_coordinates.xlsx')
reference_sheet = reference_book.active

recent_book = load_workbook('Recent_data.xlsx')
recent_sheet = recent_book['English_Date']

# Copying header row from latest data sheet

headings = [recent_sheet.cell(row=1,column=p).value for p in range(3,recent_sheet.max_column+1)]
target_col = reference_sheet.max_column+1

for index,value in enumerate(headings):
    reference_sheet.cell(row=1,column=target_col+index,value=value)



for a in range(2,reference_sheet.max_row+1):
    searching_station = reference_sheet.cell(row=a,column=2).value

    for x in range(2,recent_sheet.max_row+1):
        found_station = recent_sheet.cell(row=x,column=2).value

        if str(searching_station).lower()==str(found_station).lower():
            found_items = [recent_sheet.cell(row=x,column=u).value for u in range(3,recent_sheet.max_column+1)]

            for index,value in enumerate(found_items):
                reference_sheet.cell(row=a, column=target_col + index, value=value)


reference_book.save('01_STW_updated_with_recent_data.xlsx')

