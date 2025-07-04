import openpyxl
from openpyxl import Workbook,load_workbook

wb = load_workbook('2_formatted.xlsx')
new_wb_STW = Workbook()
new_wb_DTW = Workbook()

for sheet in wb.worksheets:
    if 'sheet' in str(sheet.title).lower():
        continue
    else:
        current_district_name = sheet.title

        new_ws_STW = new_wb_STW.create_sheet(title=f"{sheet.title}_STW")
        new_ws_DTW = new_wb_DTW.create_sheet(title=f"{sheet.title}_DTW")
        header_data = [sheet.cell(row=1,column=k).value for k in range(1,sheet.max_column+1)]

        header_data[0]='District'
        header_data[1]='Station_Name'

        new_ws_DTW.append(header_data)
        new_ws_STW.append(header_data)

        for i in range (2,sheet.max_row+1):
            value_in_first_column = sheet.cell(row=i,column=1).value
            if "DTW" in str(value_in_first_column):
                DTW_Data = [sheet.cell(row=i,column=j).value for j in range(1,sheet.max_column+1)]
                DTW_Data[0]= current_district_name
                new_ws_DTW.append(DTW_Data)

            else:
                STW_Data = [sheet.cell(row=i,column=j).value for j in range(1,sheet.max_column+1)]
                STW_Data[0]= current_district_name
                new_ws_STW.append(STW_Data)

new_wb_DTW.save("3_DTW_formatted_from_2011.xlsx")
new_wb_STW.save("3_STW_formatted_from_2011.xlsx")

        




