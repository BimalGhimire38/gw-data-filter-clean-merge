import openpyxl
from openpyxl import Workbook,load_workbook
import pandas,datetime

wb = load_workbook('WL_Data_SMJ_2011_2018.xlsx')

# for sheet in wb.worksheets:
#     print(sheet.title)
#     print(sheet.max_row,sheet.max_column)
new_wb = Workbook()

def pre_formatiing(wb,new_wb):

    """
    This function formats the data in the given workbook to make it easier to read and process for machine learning models. 
    
    The function works by iterating through each sheet in the workbook, copying the data to a new sheet, and then rearranging the data in the new sheet. 
    
    The data is rearranged by taking the values from row 3 of the original sheet and spreading them out horizontally across the new sheet. The values are moved in blocks of 12 columns at a time, with each block shifted 12 columns to the right. The values are also modified to remove the month and day from the date string. The headers in the first row are also modified to remove the year from the date string. 
    
    The function then saves the new workbook with the modified data to a new file. 
    """
    for sheet in wb.worksheets:
        new_ws = new_wb.create_sheet(title=sheet.title)
        for i in range (1,sheet.max_row+1):
            for j in range(1,sheet.max_column+1):
                new_ws.cell(row=i,column=j).value = sheet.cell(row=i,column=j).value
        row_start = 3
        column_start = 5
        count = 0
        for j in range(column_start+1,sheet.max_column+1):
            if j<=column_start+11:
                print(f"working on row {row_start} and column {j} with column_start {column_start}")
                # new_ws.cell(row=row_start,column=j).value = sheet.cell(row=row_start,column=j-column_start+count*12).value
                new_ws.cell(row=row_start,column=j).value = sheet.cell(row=row_start,column=j-(j-column_start)).value
            elif j==column_start+12:
                print(f"starting column shifted to {j}")
                column_start = j
                count+=1
            else:
                pass
        new_ws.delete_rows(1,2)
        for j in range(5,new_ws.max_column+1):
            current_value = new_ws.cell(row=1,column=j).value
            modified_year_value = current_value[7:]
            new_ws.cell(row=1,column=j).value = modified_year_value
        new_wb.save('1_pre_formatted.xlsx')


def Final_datetime_formatting(new_wb):
    for new_sheet in new_wb.worksheets:
        for j in range(5,new_sheet.max_column+1):
            current_year = new_sheet.cell(row=1,column=j).value
            current_month = new_sheet.cell(row=2,column=j).value
            year_month = f"{current_year}_{current_month}"
            formatted_date = datetime.datetime.strptime(year_month,"%Y_%b")
            new_sheet.cell(row=1,column=j).value = formatted_date
        new_sheet.delete_rows(2,1)
        new_wb.save('2_formatted.xlsx')
        
    
pre_formatiing(wb,new_wb)
Final_datetime_formatting(new_wb)
   



