# 2081/11/13 02:32 PM , CIT @ Pulchowk
import openpyxl
from openpyxl import Workbook,load_workbook

original_swb = load_workbook("3_STW_formatted_from_2011.xlsx")
original_dwb = load_workbook("3_DTW_formatted_from_2011.xlsx")


supp_swb = load_workbook("formatted_STW_supp_data.xlsx")
supp_dwb = load_workbook("formatted_DTW_supp_data.xlsx")

supp_sws = supp_swb.active
supp_dws = supp_dwb.active

# for Shallow Tube Wells
new_swb = Workbook()


# header_keys = [original_swb["Sunsari_STW"].cell(row=1,column=o).value for o in range(1,original_swb["Sunsari_STW"].max_column+1)]
# print(header_keys)
for current_sheet in original_swb.worksheets:
    if "STW" in current_sheet.title:
        new_sheet =new_swb.create_sheet(current_sheet.title)
        for r in range(2,current_sheet.max_row+1):
            searching_district = current_sheet.cell(row=r,column=1).value
            searching_station = current_sheet.cell(row=r,column=2).value
            for s in range(2,supp_sws.max_row+1):
                if searching_district == supp_sws.cell(row=s,column=1).value and searching_station == supp_sws.cell(row=s,column=2).value:
                    for c in range(1,current_sheet.max_column+1):
                        current_cell_value = current_sheet.cell(row=r,column=c).value
                        if current_cell_value == None or current_cell_value == "" or current_cell_value == " " or current_cell_value == "-":
                            current_sheet.cell(row=r,column=c).value = supp_sws.cell(row=s,column=c).value
        for i in range (1,current_sheet.max_row+1):
            for j in range(1,current_sheet.max_column+1):
                new_sheet.cell(row=i,column=j).value = current_sheet.cell(row=i,column=j).value

original_swb.save("1_STW_from_2011.xlsx")
# For Deep Tube Wells
new_dwb = Workbook()

for current_sheet in original_dwb.worksheets:
    if "DTW" in current_sheet.title:
        new_sheet =new_dwb.create_sheet(current_sheet.title)
        for r in range(2,current_sheet.max_row+1):
            searching_district = current_sheet.cell(row=r,column=1).value
            searching_station = current_sheet.cell(row=r,column=2).value
            for s in range(2,supp_dws.max_row+1):
                if searching_district == supp_dws.cell(row=s,column=1).value and searching_station == supp_sws.cell(row=s,column=2).value:
                    for c in range(1,current_sheet.max_column+1):
                        current_cell_value = current_sheet.cell(row=r,column=c).value
                        if current_cell_value == None or current_cell_value == "" or current_cell_value == " " or current_cell_value == "-":
                            current_sheet.cell(row=r,column=c).value = supp_dws.cell(row=s,column=c).value
        for i in range (1,current_sheet.max_row+1):
            for j in range(1,current_sheet.max_column+1):
                new_sheet.cell(row=i,column=j).value = current_sheet.cell(row=i,column=j).value

original_dwb.save("1_DTW_from_2011.xlsx")

    


