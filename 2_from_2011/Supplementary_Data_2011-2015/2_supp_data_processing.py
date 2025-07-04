#2081/11/08  06:47 PM @ Home,Bungmati,Lalitpur


import openpyxl
from openpyxl import load_workbook,Workbook
import os
from datetime import datetime


swb=load_workbook("1_STW_NEW.xlsx")
dwb= load_workbook("1_DTW_NEW.xlsx")


sws = swb.active
dts = dwb.active
Jhapa_data = {}
Morang_data = {}
Sunsari_data = {}
Udayapur_data = {}

new_swb = Workbook()
new_dwb = Workbook()

new_sws = new_swb.active
new_sws.title = "STW_Supplementary_data_(11_55)"

new_dws = new_dwb.active
new_dws.title = "DTW_Supplementary_data_(11_55)"

# For STWs

header_keys = [sws.cell(row=1,column=c).value for c in range(1,sws.max_column+1)]
print(header_keys)
for i in range(2,sws.max_row+1):
    print(f"Working on row no {i}")
    district = sws.cell(row=i,column=1).value
    current_station = sws.cell(row=i,column=2).value
    # district == "Jhapa":
    dict_name = globals().get(f"{district}_data",None)
    current_data = [sws.cell(row=i,column=j).value for j in range(1,sws.max_column+1)]
    current_data_dict ={}
    for i in range(0,len(header_keys)):
        current_data_dict[header_keys[i]]=current_data[i]
    if current_station not in dict_name:
        dict_name[current_station]=current_data_dict
    else:
        dict_name[current_station]={k:dict_name[current_station].get(k) if dict_name[current_station].get(k) is not None else current_data_dict.get(k) for k in dict_name[current_station]}
# Now writing the data into excel sheet
for districts in ['Jhapa','Morang','Sunsari','Udayapur']:
    current_dictonary = globals().get(f"{districts}_data",None)
    if districts=='Jhapa':
        # headers = list(current_dictonary[0].keys())
        new_sws.append(header_keys)
        for each_station in current_dictonary:
            station_data_list =list(current_dictonary[each_station].values())
            new_sws.append(station_data_list)

    else:
        for each_station in current_dictonary:
            station_data_list =list(current_dictonary[each_station].values())
            new_sws.append(station_data_list)
new_swb.save("formatted_STW_supp_data.xlsx")


# For DTWs
DJhapa_data = {}
DMorang_data = {}
DSunsari_data = {}
DUdayapur_data = {}
Dheader_keys = [dts.cell(row=1,column=c).value for c in range(1,dts.max_column+1)]
for i in range(2,dts.max_row+1):
    print(f"Working on row no {i}")
    district = dts.cell(row=i,column=1).value
    current_station = dts.cell(row=i,column=2).value
    # district == "Jhapa":
    dict_name = globals().get(f"D{district}_data",None)
    current_data = [dts.cell(row=i,column=j).value for j in range(1,dts.max_column+1)]
    current_data_dict ={}
    for i in range(0,len(header_keys)):
        current_data_dict[header_keys[i]]=current_data[i]
    if current_station not in dict_name:
        dict_name[current_station]=current_data_dict
    else:
        dict_name[current_station]={k:dict_name[current_station].get(k) if dict_name[current_station].get(k) is not None else current_data_dict.get(k) for k in dict_name[current_station]}

# Now writing the data into excel sheet
for districts in ['Jhapa','Morang','Sunsari','Udayapur']:
    current_dictonary = globals().get(f"D{districts}_data",None)
    if districts=='Jhapa':
        # headers = list(current_dictonary[0].keys())
        new_dws.append(Dheader_keys)
        for each_station in current_dictonary:
            station_data_list =list(current_dictonary[each_station].values())
            new_dws.append(station_data_list)

    else:
        for each_station in current_dictonary:
            station_data_list =list(current_dictonary[each_station].values())
            new_dws.append(station_data_list)
new_dwb.save("formatted_DTW_supp_data.xlsx")







































