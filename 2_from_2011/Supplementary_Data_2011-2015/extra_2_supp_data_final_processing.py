#2081/11/08  02:45 PM @ IOE,Pulchowk Campus CIT Hall

import openpyxl
from openpyxl import load_workbook,Workbook
import os
from datetime import datetime
from collections import defaultdict

swb=load_workbook("1_STW_NEW.xlsx")
dwb= load_workbook("1_DTW_NEW.xlsx")


sws = swb.active
dtw = dwb.active

stw_stations = {}
for i in range(1,sws.max_row+1):
    district = sws.cell(row=i,column=1).value
    station = sws.cell(row=i,column=2).value
    if district in stw_stations:
        stw_stations[district].append(station)
    else:
        stw_stations[district]=[station]
print(stw_stations)












