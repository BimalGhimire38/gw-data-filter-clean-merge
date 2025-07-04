#2081/11/06  12:31 PM @ IOE,Pulchowk Campus CIT Hall

import openpyxl
from openpyxl import load_workbook,Workbook
import os
from datetime import datetime

def date_formatter(year_value):
    month_list = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    new_date = [datetime.strptime(f"01-{a_month}-{year_value}", "%d-%b-%Y").strftime("%Y-%m-%d %H:%M:%S") for a_month in month_list]
    return new_date

files = os.listdir()
# xlsx_files = [file for file in files if file.endswith(".xlsx")]
xlsx_files = [file for file in files if file.endswith(".xlsx") and file not in ("STW_NEW.xlsx", "DTW_NEW.xlsx")]

# print(xlsx_files)
header_dates =[]
for year in range(2011,2016):
    new_dates = date_formatter(year)
    for a_date in new_dates:
        header_dates.append(a_date)
name_header = ['District','Station_Name','X','Y']

header_all = name_header + header_dates
# Converting to dictonary 
dict_header_all = {key: None for key in header_all}

new_swb = Workbook()
new_stw_ws = new_swb.active
new_stw_ws.append(header_all)
new_stw_ws.title = "STW_Supplementry_Data"

new_dwb = Workbook()
new_dtw_ws = new_dwb.active
new_dtw_ws.append(header_all)
new_dtw_ws.title = "DTW_Supplementry_Data"

for file in xlsx_files:
    print(f"Processing {file}")
    wb = load_workbook(file)
    ws = wb.active
    STW_counter =0
    DTW_counter =0
    current_district = str(file).split("_")[0]
    current_year = str(file).split("_")[1].split(".")[0]
    # print(current_district,current_year)
    for i in range(1,ws.max_row+1):
        well_identifier = ws.cell(i,2).value
        if "STW" in str(well_identifier):
            dict_of_data = dict_header_all.copy()
            # print(f"{well_identifier} is a STW")
            STW_counter+=1
            # Data Extraction 
            initial_data_found = [ws.cell(row=i,column=col).value for col in range(3,6)]
            initial_data_found.insert(0,current_district)
            for key, value in zip(list(dict_of_data.keys())[:len(initial_data_found)], initial_data_found):
                dict_of_data[key] = value
            # now this initial_data_found list contains district,stn name , X, Y
            # Data Searching on the basis of date from new workbook's header.

            # for a_date in header_dates:
                # Dates start from index 3, and column 4 
                # formatting the date in the workbook for easy comparision
            # Creating a new dictonary to store the date and value only
            value_only_dict = {}
            for k in range(6,18):
                month_short_form = ws.cell(row=1,column=k).value
                formatted_date= datetime.strptime(f"01-{month_short_form}-{current_year}", "%d-%b-%Y").strftime("%Y-%m-%d %H:%M:%S")
                if month_short_form:  # Ensures it's not None
                    formatted_date = datetime.strptime(f"01-{month_short_form}-{current_year}", "%d-%b-%Y").strftime("%Y-%m-%d %H:%M:%S")
                else:
                    print(f"Warning: Empty month in file {file} at column {k}")
                    formatted_date = None  # Handle missing months properly
                gw_value = ws.cell(row=i,column=k).value
                value_only_dict[formatted_date] = gw_value
            dict_of_data.update(value_only_dict)
            new_stw_ws.append(list(dict_of_data.values()))
        elif "DTW" in str(well_identifier):
            # print(f"{well_identifier} is a DTW")
            DTW_counter+=1
            dict_of_data = dict_header_all.copy()
            # Data Extraction 
            initial_data_found = [ws.cell(row=i,column=col).value for col in range(3,6)]
            initial_data_found.insert(0,current_district)
            for key, value in zip(list(dict_of_data.keys())[:len(initial_data_found)], initial_data_found):
                dict_of_data[key] = value
            # now this initial_data_found list contains district,stn name , X, Y
            # Data Searching on the basis of date from new workbook's header.

            # for a_date in header_dates:
                # Dates start from index 3, and column 4 
                # formatting the date in the workbook for easy comparision
            # Creating a new dictonary to store the date and value only
            value_only_dict = {}
            for k in range(6,18):
                month_short_form = ws.cell(row=1,column=k).value
                formatted_date= datetime.strptime(f"01-{month_short_form}-{current_year}", "%d-%b-%Y").strftime("%Y-%m-%d %H:%M:%S")
                gw_value = ws.cell(row=i,column=k).value
                value_only_dict[formatted_date] = gw_value
            dict_of_data.update(value_only_dict)
            new_dtw_ws.append(list(dict_of_data.values()))
        else:
            pass

    print(f"Total {STW_counter} STW in {current_district}_{current_year}")
    print(f"Total {DTW_counter} DTW in {current_district}_{current_year}")
new_dwb.save("1_DTW_NEW.xlsx")
new_swb.save("1_STW_NEW.xlsx")


